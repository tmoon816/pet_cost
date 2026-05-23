"""客户批量导入：xlsx 解析 + 校验 + 落库。

策略：分两步——preview 跑一次 dry-run 把所有错误列出来，让用户改了再 confirm。
phone 撞库视为错误（不悄覆盖、不静默跳过），让用户决定。
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..models import Customer

TEMPLATE_HEADERS = ["姓名*", "手机号", "备注"]
TEMPLATE_HINT = (
    "字段说明：\n"
    "- 姓名*：必填，1~50 字符\n"
    "- 手机号：选填，填写时必须是 11 位数字\n"
    "- 备注：选填，最长 500 字符\n"
    "- 同一文件内手机号不能重复，已存在系统中的手机号会报错"
)


def build_template() -> bytes:
    """生成空白模板，第一行是 header，第二行起空。最后一个 sheet 写说明。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "客户"
    ws.append(TEMPLATE_HEADERS)
    ws.append(["张三", "13800000000", "示例：可删"])

    hint_ws = wb.create_sheet("填写说明")
    for line in TEMPLATE_HINT.split("\n"):
        hint_ws.append([line])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _norm(v: Any) -> str:
    """openpyxl 单元格值统一转 str trim，None / 空格全归空串。"""
    if v is None:
        return ""
    return str(v).strip()


def _parse_rows(file_bytes: bytes) -> tuple[list[dict], list[dict]]:
    """读 xlsx 第一个 sheet。返回 (rows, parse_errors)。

    rows: [{row_no, name, phone, note}]，row_no 从 2 开始（跳过 header）
    parse_errors: 文件级错误（无 header / 无数据等）
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        return [], [{"row": 0, "message": f"无法读取文件：{e}"}]

    ws = wb.worksheets[0] if wb.worksheets else None
    if ws is None:
        return [], [{"row": 0, "message": "文件中没有任何工作表"}]

    rows: list[dict] = []
    iterator = ws.iter_rows(values_only=True)
    try:
        header = next(iterator)
    except StopIteration:
        return [], [{"row": 0, "message": "文件为空"}]

    if not header or _norm(header[0]) not in {"姓名*", "姓名"}:
        return [], [{"row": 1, "message": "首行必须是模板表头（姓名* / 手机号 / 备注）"}]

    for idx, raw in enumerate(iterator, start=2):
        if not raw or all(_norm(c) == "" for c in raw):
            continue
        name = _norm(raw[0]) if len(raw) > 0 else ""
        phone = _norm(raw[1]) if len(raw) > 1 else ""
        note = _norm(raw[2]) if len(raw) > 2 else ""
        rows.append({"row_no": idx, "name": name, "phone": phone, "note": note})
    return rows, []


def _validate_row(row: dict) -> list[str]:
    errs: list[str] = []
    name = row["name"]
    phone = row["phone"]
    note = row["note"]

    if not name:
        errs.append("姓名不能为空")
    elif len(name) > 50:
        errs.append("姓名超过 50 字符")

    if phone:
        if not (phone.isdigit() and len(phone) == 11):
            errs.append("手机号必须是 11 位数字")

    if note and len(note) > 500:
        errs.append("备注超过 500 字符")
    return errs


def preview(db: Session, file_bytes: bytes) -> dict:
    """dry-run。返回 {ok, errors, total}。绝不写库。"""
    rows, parse_errs = _parse_rows(file_bytes)
    if parse_errs:
        return {"ok": [], "errors": parse_errs, "total": 0}

    if not rows:
        return {"ok": [], "errors": [{"row": 0, "message": "没有数据行"}], "total": 0}

    # 收集所有非空 phone，一次性查 DB
    phones_in_file = [r["phone"] for r in rows if r["phone"]]
    db_phones: set[str] = set()
    if phones_in_file:
        db_phones = set(
            db.scalars(
                select(Customer.phone).where(Customer.phone.in_(phones_in_file))
            ).all()
        )

    # 文件内重复 phone 检测：phone → 第一次出现的 row_no
    seen_phone: dict[str, int] = {}

    ok: list[dict] = []
    errors: list[dict] = []
    for r in rows:
        msgs = _validate_row(r)

        if r["phone"]:
            if r["phone"] in db_phones:
                msgs.append(f"手机号 {r['phone']} 已存在")
            elif r["phone"] in seen_phone:
                msgs.append(f"手机号 {r['phone']} 与第 {seen_phone[r['phone']]} 行重复")
            else:
                seen_phone[r["phone"]] = r["row_no"]

        if msgs:
            errors.append({"row": r["row_no"], "message": "；".join(msgs), "name": r["name"]})
        else:
            ok.append(r)

    return {"ok": ok, "errors": errors, "total": len(rows)}


def commit(db: Session, file_bytes: bytes) -> dict:
    """重新跑 preview，确保数据有效再写。返回 {inserted, errors}。
    若仍有错误，整批拒绝以保证原子性，避免半成品。"""
    result = preview(db, file_bytes)
    if result["errors"]:
        return {"inserted": 0, "errors": result["errors"]}

    objs: list[Customer] = []
    for r in result["ok"]:
        objs.append(
            Customer(
                name=r["name"],
                phone=r["phone"] or None,
                note=r["note"] or None,
            )
        )
    db.add_all(objs)
    db.commit()
    return {"inserted": len(objs), "errors": []}
