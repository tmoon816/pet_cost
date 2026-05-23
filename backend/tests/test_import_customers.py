"""客户批量导入测试：模板下载、preview 校验、commit 原子性。"""

from __future__ import annotations

import io

import pytest
from openpyxl import Workbook, load_workbook

from app.crud.import_customers import build_template, commit, preview
from app.models import Customer


def _make_xlsx(rows: list[list]) -> bytes:
    """rows 第一行是 header。"""
    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_template_has_headers_and_hint():
    data = build_template()
    wb = load_workbook(io.BytesIO(data), read_only=True)
    sheets = wb.sheetnames
    assert "客户" in sheets
    assert "填写说明" in sheets
    rows = list(wb["客户"].iter_rows(values_only=True))
    assert rows[0] == ("姓名*", "手机号", "备注")


def test_preview_valid_rows_pass(db_session):
    db = db_session()
    try:
        data = _make_xlsx([
            ["姓名*", "手机号", "备注"],
            ["新客户A", "13800000001", "vip"],
            ["新客户B", "", ""],
        ])
        result = preview(db, data)
        assert result["total"] == 2
        assert len(result["ok"]) == 2
        assert result["errors"] == []
    finally:
        db.close()


def test_preview_phone_conflict_in_db(db_session):
    db = db_session()
    try:
        db.add(Customer(name="老王", phone="13800000001"))
        db.commit()

        data = _make_xlsx([
            ["姓名*", "手机号", "备注"],
            ["新客户", "13800000001", ""],
        ])
        result = preview(db, data)
        assert len(result["errors"]) == 1
        assert "已存在" in result["errors"][0]["message"]
    finally:
        db.close()


def test_preview_phone_dup_in_file(db_session):
    db = db_session()
    try:
        data = _make_xlsx([
            ["姓名*", "手机号", "备注"],
            ["A", "13800000001", ""],
            ["B", "13800000001", ""],
        ])
        result = preview(db, data)
        assert len(result["errors"]) == 1
        assert "重复" in result["errors"][0]["message"]
        assert result["errors"][0]["row"] == 3
    finally:
        db.close()


def test_preview_invalid_phone_format(db_session):
    db = db_session()
    try:
        data = _make_xlsx([
            ["姓名*", "手机号", "备注"],
            ["A", "abc", ""],
            ["B", "138", ""],
        ])
        result = preview(db, data)
        assert len(result["errors"]) == 2
        for e in result["errors"]:
            assert "11 位数字" in e["message"]
    finally:
        db.close()


def test_preview_missing_name(db_session):
    db = db_session()
    try:
        data = _make_xlsx([
            ["姓名*", "手机号", "备注"],
            ["", "13800000001", ""],
        ])
        result = preview(db, data)
        assert len(result["errors"]) == 1
        assert "姓名" in result["errors"][0]["message"]
    finally:
        db.close()


def test_preview_bad_header(db_session):
    db = db_session()
    try:
        data = _make_xlsx([["foo", "bar"], ["x", "y"]])
        result = preview(db, data)
        assert len(result["errors"]) == 1
        assert result["errors"][0]["row"] == 1
    finally:
        db.close()


def test_preview_empty_file(db_session):
    db = db_session()
    try:
        data = _make_xlsx([["姓名*", "手机号", "备注"]])
        result = preview(db, data)
        assert result["errors"] != []
    finally:
        db.close()


def test_commit_atomic_rejects_when_any_error(db_session):
    db = db_session()
    try:
        data = _make_xlsx([
            ["姓名*", "手机号", "备注"],
            ["A", "13800000001", ""],
            ["B", "abc", ""],  # 一个非法手机
        ])
        result = commit(db, data)
        assert result["inserted"] == 0
        assert result["errors"] != []

        # 数据库未写入
        count = db.query(Customer).count()
        assert count == 0
    finally:
        db.close()


def test_commit_writes_when_all_valid(db_session):
    db = db_session()
    try:
        data = _make_xlsx([
            ["姓名*", "手机号", "备注"],
            ["新客户A", "13800000001", "备注1"],
            ["新客户B", "", ""],
        ])
        result = commit(db, data)
        assert result["inserted"] == 2
        assert result["errors"] == []

        rows = db.query(Customer).order_by(Customer.id).all()
        assert len(rows) == 2
        assert rows[0].name == "新客户A"
        assert rows[0].phone == "13800000001"
        assert rows[1].phone is None
    finally:
        db.close()
